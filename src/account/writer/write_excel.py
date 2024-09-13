from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from account.utils.utils import get_parent_dir
from account.writer.base_writer import BaseWriter


class XlsWriter(BaseWriter):
    def __init__(self, base_path):
        super().__init__(base_path)

        parent_dir = get_parent_dir(__file__)
        self.template = f"{parent_dir}\\config\\dues_template.xlsx"

    def write(self, df, file_nm):
        # Load your existing Excel file (template)
        book = load_workbook(self.template)

        xls_path = f"{self.path}\\{file_nm}"

        # Select the desired sheet where you want to write the data
        writer = pd.ExcelWriter(xls_path, engine="openpyxl")
        writer._book = book
        sheet_name = "Dues"

        # Write the DataFrame to the existing sheet starting from a specific cell
        # Example: starting from cell B2
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=3,
            startcol=0,
            index=False,
            header=False,
        )

        # Save the changes
        writer.close()

        self.__format_dues(df, book)

        book.save(xls_path)

        return xls_path

    def __format_dues(self, df, book):
        ws = book.worksheets[0]
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        accounting_format = '_(₹* #,##0.00_);_(₹* (#,##0.00);_(₹* "-"??_);_(@_)'
        for row in range(1, len(df) + 6):
            for col in ["A", "B", "C", "D", "E", "F", "G"]:
                cell = f"{col}{row}"
                ws[cell].border = thin_border

                if col not in ["A", "B"]:
                    ws[cell].number_format = accounting_format
        ws[cell] = df["Total Dues"].sum()
        cur_date = datetime.now().strftime("%b %d, %Y")
        ws["A2"] = f"Dues as of {cur_date}"
