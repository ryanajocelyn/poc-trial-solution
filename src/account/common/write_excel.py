from datetime import datetime

import pandas as pd
from PyPDF2 import PdfFileReader, PdfFileWriter, PdfWriter, PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from account.utils.utils import get_parent_dir


class XlsWriter:
    def __init__(self, base_path):
        parent_dir = get_parent_dir(__file__)
        self.template = f"{parent_dir}\\config\\dues_template.xlsx"
        self.path = f"{base_path}\\Dues"

    def write(self, df):
        # Load your existing Excel file (template)
        book = load_workbook(self.template)

        xls_path = f"{self.path}\\dues.xlsx"

        # Select the desired sheet where you want to write the data
        writer = pd.ExcelWriter(xls_path, engine="openpyxl")
        writer._book = book
        sheet_name = "Dues"

        # Write the DataFrame to the existing sheet starting from a specific cell
        # Example: starting from cell B2
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=3,
            startcol=0,
            index=False,
            header=False,
        )

        # Save the changes
        writer._save()

        ws = book.worksheets[0]
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        accounting_format = '_(₹* #,##0.00_);_(₹* (#,##0.00);_(₹* "-"??_);_(@_)'

        for row in range(1, len(df) + 6):
            for col in ["A", "B", "C", "D", "E", "F", "G"]:
                cell = f"{col}{row}"
                ws[cell].border = thin_border

                if col not in ["A", "B"]:
                    ws[cell].number_format = accounting_format

        ws[cell] = df["Total Dues"].sum()
        cur_date = datetime.now().strftime("%b %d, %Y")
        ws["A2"] = f"Dues as of {cur_date}"
        book.save(xls_path)

        # ws.sheet_state = "hidden"
        tmp_pdf = f"{self.path}\\dues_tmp.pdf"
        book.save(tmp_pdf)

        # ws.sheet_state = "visible"

        # Create a PDF writer
        pdf_writer = PdfWriter()

        # Merge the temporary PDF into the final PDF
        pdf_reader = PdfReader(open(tmp_pdf, "rb"))
        pdf_writer.addPage(pdf_reader.getPage(0))

        # Write the final PDF
        pdf_path = f"{self.path}\\dues_pdf.pdf"
        with open(pdf_path, "wb") as f:
            pdf_writer.write(f)
